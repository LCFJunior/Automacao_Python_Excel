# Import

import openpyxl
from openpyxl.styles import Font
from babel.numbers import format_currency
from openpyxl.styles import Font


# Carregar o arquivo Excel

book = openpyxl.load_workbook('Vendas.xlsx')


# Obter as páginas relevantes

Vendas_Pag = book['Vendas']
Pagamentos_Pag = book['Pagamentos']
Calc_Comissoes_Pag = book.create_sheet('Calculo de Comissoes', 1)
Validacao_Pag = book.create_sheet('Validacao de Pagamentos', 2)


# Inicializar listas para armazenar dados

Nome_Coluna_Pag1 = []
Valor_Coluna_Pag1 = []
Ajuste_Coluna_Pag1 = []
Pagamentos_Incorretos = []


# Implementação do nome dos vendedores - Tarefa 1

for celulas in Vendas_Pag[1]:
    if celulas.value == 'Nome do Vendedor':
        col_index_nome_venda = celulas.column
        break
else:
    raise ValueError('Cabeçalho "Nome do Vendedor" não existe!')

for linhas in Vendas_Pag.iter_rows(min_col=col_index_nome_venda, max_col=col_index_nome_venda, min_row=2):
    for celulas in linhas:
        Nome_Coluna_Pag1.append(celulas.value)

Calc_Comissoes_Pag['A1'] = 'Nome do Vendedor'
Calc_Comissoes_Pag['A1'].font = Font(bold=True)
Calc_Comissoes_Pag.column_dimensions['A'].width = 20

for i, valor in enumerate(Nome_Coluna_Pag1, start=2):
    Calc_Comissoes_Pag[f'A{i}'] = valor


# Implementação de comissão bruta (Sem ajuste de regras estabelecidas) - Tarefa 1

for celulas in Vendas_Pag[1]:
    if celulas.value == 'Valor da Venda':
        col_index_valor_venda = celulas.column
        break
else:
    raise ValueError('Cabeçalho "Valor da Venda" não existe!')

for linhas in Vendas_Pag.iter_rows(min_col=col_index_valor_venda, max_col=col_index_valor_venda, min_row=2):
    for celulas in linhas:
        valor = celulas.value
        if isinstance(valor, (int, float)):
            Valor_Coluna_Pag1.append(valor)
        elif isinstance(valor, str) and valor.replace('.', '', 1).isdigit():
            Valor_Coluna_Pag1.append(float(valor))
        else:
            Valor_Coluna_Pag1.append(None)

Calc_Comissoes_Pag['B1'] = 'Valor da Comissao'
Calc_Comissoes_Pag['B1'].font = Font(bold=True)
Calc_Comissoes_Pag.column_dimensions['B'].width = 20

for i, valor in enumerate(Valor_Coluna_Pag1, start=2):
    if valor is not None:
        Valor_Formatado = Calc_Comissoes_Pag[f'B{i}'] = format_currency(valor * 0.1 if valor != 0 else 0, 'BRL', locale='pt_BR')
        Calc_Comissoes_Pag[f'B{i}'] = Valor_Formatado


# Implementação de comissão após ajuste (Seguindo regras estabelecidas) - Tarefa 1

for celulas in Vendas_Pag[1]:
    if celulas.value == 'Valor da Venda':
        col_index_valor_venda = celulas.column
    elif celulas.value == 'Canal de Venda':
        col_index_canal_venda = celulas.column
        break
else:
    raise ValueError('Cabeçalho "Valor da Venda" ou "Canal de Venda" não existe!')

for linha in Vendas_Pag.iter_rows(min_row=2):
    valor_venda = linha[col_index_valor_venda - 1].value
    canal_venda = linha[col_index_canal_venda - 1].value
    if isinstance(valor_venda, (int, float)):
        comissao_vendedor = valor_venda * 0.1
        if canal_venda == 'Online' and valor_venda >= 1500:
            comissao_vendedor -= comissao_vendedor * 0.3
        elif canal_venda == 'Online':
            comissao_vendedor -= comissao_vendedor * 0.2
        elif valor_venda >= 1500:
            comissao_gerente = comissao_vendedor * 0.1
            comissao_vendedor -= comissao_gerente
        Ajuste_Coluna_Pag1.append(comissao_vendedor)

Calc_Comissoes_Pag['C1'] = 'Comissao Ajustada'
Calc_Comissoes_Pag['C1'].font = Font(bold=True)
Calc_Comissoes_Pag.column_dimensions['C'].width = 20

for i, valor in enumerate(Ajuste_Coluna_Pag1, start=2):
    if valor is not None:
        Valor_Formatado = Calc_Comissoes_Pag[f'C{i}'] = format_currency(valor, 'BRL', locale='pt_BR')


# Comparação de valores corretos e incorretos - Tarefa 2

for celulas in Pagamentos_Pag[1]:
    if celulas.value == 'Comissão':
        col_index_comissao_pagamento = celulas.column
        break
else:
    raise ValueError('Cabeçalho "Comissão" ou "Nome do Vendedor" não existe!')

for celulas in Pagamentos_Pag[1]:
    if celulas.value == 'Nome do Vendedor':
        col_index_nome_pagamento = celulas.column
        break
else:
    raise ValueError('Cabeçalho "Comissão" ou "Nome do Vendedor" não existe!')

for linha in Pagamentos_Pag.iter_rows(min_row=2, values_only=True):
    nome_vendedor = linha[col_index_nome_pagamento - 1]
    valor_pago = linha[col_index_comissao_pagamento - 1]

    valor_calculado = None
    for i, nome in enumerate(Nome_Coluna_Pag1):
        if nome == nome_vendedor:
            valor_calculado = Ajuste_Coluna_Pag1[i]
            break

    if valor_calculado is not None and valor_pago != valor_calculado:
        Pagamentos_Incorretos.append({
            'Vendedor': nome_vendedor,
            'Valor Pago Incorreto': valor_pago,
            'Valor Calculado Correto': valor_calculado
        })

Validacao_Pag['A1'] = 'Vendedor'
Validacao_Pag['B1'] = 'Valor Pago Incorreto'
Validacao_Pag['C1'] = 'Valor Calculado Correto'

Validacao_Pag['A1'].font = Font(bold=True)
Validacao_Pag['B1'].font = Font(bold=True)
Validacao_Pag['C1'].font = Font(bold=True)

Validacao_Pag.column_dimensions['A'].width = 30
Validacao_Pag.column_dimensions['B'].width = 30
Validacao_Pag.column_dimensions['C'].width = 30

for i, pagamento in enumerate(Pagamentos_Incorretos, start=2):
    Validacao_Pag[f'A{i}'] = pagamento['Vendedor']
    Validacao_Pag[f'B{i}'] = format_currency(pagamento['Valor Pago Incorreto'], 'BRL', locale='pt_BR')
    Validacao_Pag[f'C{i}'] = format_currency(pagamento['Valor Calculado Correto'], 'BRL', locale='pt_BR')


# Salvar alterações na planilha

book.save('Vendas_V2.xlsx')
